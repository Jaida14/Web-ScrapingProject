
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2023/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##

table_rows = soup.findAll("tr") #each row is tr, but each elemenet in that one row is td 
#next(table_rows)

'''for row in table_rows[1:6]: #the third row with the states all the way to all the states
    td = row.findAll("td") #you have to do tr on the outside and then break it down under the for loop
   
    
    print(f'Rank: {td[0].text}')
    print(f'Release: {td[1].text}')
    print(f'Gross: {td[5].text}')
    print(f'Theaters: {td[6].text}')
    print(f'Total Gross: {td[7].text}')
    print(f'Release Date: {td[8].text}')
    print(f'Distributor: {td[9].text}')      
    print(" ")''' #this works so we can write to the excel sheet

wb = xl.workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'Rank'
ws['B1'] = 'Release'
ws['C1'] = 'Gross'
ws['D1'] = 'Theaters'
ws['E1'] = 'Avg. Gross / Theater'

for x in range(1,6): #the third row with the states all the way to all the states
    td = table_rows[x].findAll("td") #you have to do tr on the outside and then break it down under the for loop
   
    
    rank = td[0].text
    release = td[1].text
    gross = int(td[5].text.replace('$','').replace(',', '')) #need to strip the numbers to be just the integers
    theaters = int(td[6].text.replace(',', ''))
    
    avg = gross / theaters

    ws['A' + str(x + 1)] = rank #A2, A3, A4...
    ws['B' + str(x + 1)] = release
    ws['C' + str(x + 1)] = gross
    ws['D' + str(x + 1)] = theaters
    ws['E' + str(x + 1)] = avg


ws.column_dimensions['A'].width = 7
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 30

header_font = Font(size = 16, bold = True)


for cell in ws[1:1]: #for the first row only
    cell.font = header_font

for cell in ws["D:D"]:
    cell.number_format = "#,##0"

for cell in ws["C:C"]:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws["E:E"]:
    cell.number_format = u'"$ "#,##0.00'

wb.save('BoxOfficeRepot.xlsx')