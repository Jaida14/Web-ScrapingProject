import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook() #made a workbook

ws = wb.active #activated it

ws.title = 'First Sheet' #titled the first sheet (auto first sheet of workbook)

wb.create_sheet(index=1, title = 'Second Sheet') #made a second sheet of the workbook


#anything you want to write to the workbook, you need to specify to what worksheet
ws['A1'] = "Invoice"
ws['A2'] = "Tires"
ws['A3'] = "Breaks"
ws['A4'] = "Alignment"


#method 1
#ws['A1'].font = Font(name = "Times New Roman", size = 24, bold = True) #this will only affect 'Invoice'

#method 2
header_font = Font(name = "Times New Roman", size = 24, bold = True)

ws['A1'].font = header_font

#merge cells
ws.merge_cells('A1:B1')

#if you want to unmerge
#ws.unmerge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = Font(size = 16, bold = True)









wb.save('PythonToExcel.xlsx')